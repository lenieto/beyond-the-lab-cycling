"""
openpose_core.py
Universidad de los Andes · Semillero EMBS
Proyecto: Análisis biomecánico en ciclismo con OpenPose
"""
 
import os, sys, json, math, time, shutil, subprocess, platform
from pathlib import Path
from datetime import datetime
import numpy as np
 
IS_WINDOWS = platform.system() == "Windows"
IS_MAC     = platform.system() == "Darwin"
 
# ── Rutas del proyecto ──────────────────────────────────────────────────────
WORKDIR    = Path(__file__).parent.parent.resolve()   # openpose-workdir/
DATA_DIR   = WORKDIR / "data" / "raw"                 # videos originales
OUTPUT_DIR = WORKDIR / "outputs"                      # resultados
 
for d in [DATA_DIR, OUTPUT_DIR]:
    d.mkdir(parents=True, exist_ok=True)
 
# ── Docker ──────────────────────────────────────────────────────────────────
DOCKER_IMAGE    = "openpose-local"
DOCKER_PLATFORM = "linux/amd64"
 
def _docker_env():
    env = os.environ.copy()
    if IS_MAC:
        # Detecta el socket de Docker Desktop automáticamente
        possible_sockets = [
            Path.home() / ".docker" / "run" / "docker.sock",
            Path("/var/run/docker.sock"),
        ]
        for s in possible_sockets:
            if s.exists():
                env["DOCKER_HOST"] = f"unix://{s}"
                break
    return env
 
def check_docker() -> tuple[bool, str]:
    try:
        result = subprocess.run(
            ["docker", "info"],
            capture_output=True, text=True, timeout=10,
            env=_docker_env()
        )
        if result.returncode == 0:
            return True, "Docker está corriendo correctamente."
        else:
            return False, "Docker no está corriendo. Abre Docker Desktop."
    except FileNotFoundError:
        return False, "Docker no está instalado."
    except subprocess.TimeoutExpired:
        return False, "Docker tardó demasiado. ¿Está abierto?"
 
def check_openpose_image() -> tuple[bool, str]:
    try:
        result = subprocess.run(
            ["docker", "image", "inspect", DOCKER_IMAGE],
            capture_output=True, text=True, timeout=10,
            env=_docker_env()
        )
        if result.returncode == 0:
            return True, f"Imagen '{DOCKER_IMAGE}' encontrada."
        else:
            return False, (
                f"La imagen Docker '{DOCKER_IMAGE}' no existe.\n"
                "Construye la imagen primero siguiendo las instrucciones del README."
            )
    except Exception as e:
        return False, str(e)
 
def get_frame_count(video_path):
    try:
        import cv2
        cap = cv2.VideoCapture(str(video_path))
        n   = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = cap.get(cv2.CAP_PROP_FPS)
        cap.release()
        return (n if n > 0 else 300), (fps if fps > 0 else 30)
    except Exception:
        return 300, 30
 
def convert_avi_to_mp4(avi_path, mp4_path):
    try:
        result = subprocess.run(
            ["ffmpeg", "-y", "-i", str(avi_path), str(mp4_path)],
            capture_output=True, text=True
        )
        return result.returncode == 0 and Path(mp4_path).exists()
    except Exception:
        return False
 
def run_openpose(video_path, output_dir, subject_id=None, progress_callback=None, status_callback=None):
    import threading
    video_path = Path(video_path)
    stem = video_path.stem.replace(" ", "_")
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"{stem}_{ts}"
 
    # Carpeta de salida para este video
    folder_name = subject_id if subject_id else stem
    run_dir  = Path(output_dir) / folder_name
    json_dir = run_dir / "json"
    run_dir.mkdir(parents=True, exist_ok=True)
    json_dir.mkdir(parents=True, exist_ok=True)
 
    label      = subject_id if subject_id else stem
    avi_path   = run_dir / f"{label}_skeleton.avi"
    mp4_path   = run_dir / "skeleton.mp4"
    excel_path = run_dir / "angles.xlsx"
    chart_path = run_dir / "charts.png"
 
    # Copia el video a data/raw/ si no está ya ahí
    dst_video = DATA_DIR / video_path.name
    if not dst_video.exists():
        shutil.copy2(str(video_path), str(dst_video))
 
    total_frames, fps = get_frame_count(video_path)
 
    if status_callback:
        status_callback("OpenPose procesando…")
 
    stop_monitor = threading.Event()
    if progress_callback:
        def _monitor():
            while not stop_monitor.is_set():
                n = len([f for f in os.listdir(str(json_dir)) if f.endswith(".json")])
                progress_callback(n, total_frames)
                time.sleep(0.8)
        threading.Thread(target=_monitor, daemon=True).start()
 
    if IS_WINDOWS:
        input_mount  = str(DATA_DIR).replace("\\", "/")
        output_mount = str(run_dir).replace("\\", "/")
        if input_mount[1] == ":":
            input_mount  = "/" + input_mount[0].lower() + input_mount[2:]
            output_mount = "/" + output_mount[0].lower() + output_mount[2:]
    else:
        input_mount  = str(DATA_DIR)
        output_mount = str(run_dir)
 
    cmd = [
        "docker", "run", "--rm",
        "--platform", DOCKER_PLATFORM,
        "-v", f"{input_mount}:/openpose/INPUT",
        "-v", f"{output_mount}:/openpose/OUTPUT",
        DOCKER_IMAGE,
        "./build/examples/openpose/openpose.bin",
        "--video", f"INPUT/{video_path.name}",
        "--write_video", f"OUTPUT/{name}_skeleton.avi",
        "--write_json", "OUTPUT/json/",
        "--display", "0"
    ]
 
    result = subprocess.run(cmd, capture_output=True, text=True, env=_docker_env())
    stop_monitor.set()
 
    if result.returncode != 0:
        return {"error": result.stderr[:800]}
 
    if progress_callback:
        progress_callback(total_frames, total_frames)
 
    if status_callback:
        status_callback("Convirtiendo video a MP4…")
    convert_avi_to_mp4(avi_path, mp4_path)
 
    if status_callback:
        status_callback("Generando Excel y gráficas…")
 
    df = None
    try:
        df = json_to_dataframe(str(json_dir), fps=fps)
        if df is not None and not df.empty:
            dataframe_to_excel(df, excel_path)
            make_angle_charts(df, chart_path, name=stem)
    except Exception as e:
        print(f"[openpose_core] Warning al generar Excel: {e}")
 
    return {
        "name":   stem,
        "mp4":    str(mp4_path)    if mp4_path.exists()   else "",
        "avi":    str(avi_path)    if avi_path.exists()   else "",
        "excel":  str(excel_path)  if excel_path.exists() else "",
        "charts": str(chart_path)  if chart_path.exists() else "",
        "json":   str(json_dir),
        "folder": str(run_dir),
        "frames": total_frames,
        "fps":    fps,
        "df":     df,
    }
 
# ── Procesamiento de keypoints ───────────────────────────────────────────────
 
def extract_point(keypoints, index):
    b = index * 3
    return (keypoints[b], keypoints[b+1], keypoints[b+2])
 
def calc_angle(p1, p2, p3):
    if any(p[2] < 0.1 for p in [p1, p2, p3]):
        return None
    v1 = (p1[0]-p2[0], p1[1]-p2[1])
    v2 = (p3[0]-p2[0], p3[1]-p2[1])
    dot = v1[0]*v2[0] + v1[1]*v2[1]
    m1  = math.sqrt(v1[0]**2 + v1[1]**2)
    m2  = math.sqrt(v2[0]**2 + v2[1]**2)
    if m1 == 0 or m2 == 0:
        return None
    return round(math.degrees(math.acos(max(-1.0, min(1.0, dot/(m1*m2))))), 2)
 
# Articulaciones relevantes para ciclismo
JOINT_MAP = {
    "Rodilla D":  (9,  10, 11),
    "Rodilla I":  (12, 13, 14),
    "Cadera D":   (1,  9,  10),
    "Cadera I":   (1,  12, 13),
    "Codo D":     (2,  3,  4),
    "Codo I":     (5,  6,  7),
    "Hombro D":   (1,  2,  3),
    "Hombro I":   (1,  5,  6),
}
 
def json_to_dataframe(json_dir, fps=30):
    try:
        import pandas as pd
    except ImportError:
        return None
    files = sorted([f for f in os.listdir(json_dir) if f.endswith("_keypoints.json")])
    if not files:
        return None
    rows = []
    for i, fname in enumerate(files):
        try:
            with open(os.path.join(json_dir, fname), encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue
        if not data.get("people"):
            continue
        kp  = data["people"][0]["pose_keypoints_2d"]
        p   = {n: extract_point(kp, n) for n in range(25)}
        row = {"Frame": i, "Tiempo (s)": round(i / fps, 3)}
        for joint_name, (a, b, c) in JOINT_MAP.items():
            row[joint_name] = calc_angle(p[a], p[b], p[c])
        rows.append(row)
    return pd.DataFrame(rows) if rows else None
 
def dataframe_to_excel(df, excel_path):
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False
    excel_path  = Path(excel_path)
    angle_cols  = [c for c in df.columns if c not in ["Frame", "Tiempo (s)"]]
    stats       = df[angle_cols].agg(["mean", "min", "max", "std"]).round(2)
    stats.index = ["Promedio", "Mínimo", "Máximo", "Desv.Std"]
    with pd.ExcelWriter(str(excel_path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Datos", index=False)
        stats.to_excel(writer, sheet_name="Estadísticas")
    wb = load_workbook(str(excel_path))
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill(fill_type="solid", fgColor="1E3A5F")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ml+4, 32)
    wb.save(str(excel_path))
    return True
 
def make_angle_charts(df, chart_path, name=""):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return False
    angle_cols = [c for c in df.columns if c not in ["Frame", "Tiempo (s)"]]
    fig, axes  = plt.subplots(2, 4, figsize=(18, 8))
    fig.patch.set_facecolor("#0D1117")
    if name:
        fig.suptitle(f"Ángulos Articulares — {name}", fontsize=13,
                     color="white", fontweight="bold", y=1.01)
    colors = ["#2563EB","#22C55E","#F59E0B","#EF4444",
              "#7C3AED","#06B6D4","#EC4899","#84CC16"]
    for ax, col, color in zip(axes.flat, angle_cols, colors):
        vals = df[col].dropna()
        t    = df.loc[vals.index, "Tiempo (s)"]
        ax.plot(t, vals, color=color, linewidth=1.5)
        ax.fill_between(t, vals, alpha=0.15, color=color)
        ax.set_facecolor("#161B22")
        ax.set_title(col, fontsize=8, color="white", pad=5)
        ax.set_xlabel("s", fontsize=7, color="#8B949E")
        ax.set_ylabel("°", fontsize=7, color="#8B949E")
        ax.tick_params(colors="#8B949E", labelsize=7)
        for spine in ax.spines.values():
            spine.set_edgecolor("#30363D")
        ax.grid(True, alpha=0.2, color="#30363D")
    plt.tight_layout()
    plt.savefig(str(chart_path), dpi=150, bbox_inches="tight", facecolor="#0D1117")
    plt.close()
    return True
 
def open_path(path):
    p = Path(path)
    target = str(p.parent) if p.is_file() else str(p)
    if IS_WINDOWS:
        os.startfile(target)
    else:
        subprocess.run(["open", target])
 
if __name__ == "__main__":
    print("=== openpose_core.py — Verificación del sistema ===\n")
    print(f"Sistema operativo: {platform.system()}")
    print(f"Carpeta de trabajo: {WORKDIR}")
    print(f"Data raw:           {DATA_DIR}")
    print(f"Outputs:            {OUTPUT_DIR}\n")
    ok, msg = check_docker()
    print(f"Docker: {'✓' if ok else '✗'} {msg}")
    if ok:
        ok2, msg2 = check_openpose_image()
        print(f"Imagen OpenPose: {'✓' if ok2 else '✗'} {msg2}")